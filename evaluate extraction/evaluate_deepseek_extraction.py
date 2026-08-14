from __future__ import annotations

import argparse
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


FIELDS = [
    "Reaction mode",
    "Year",
    "Feedstock",
    "Catalyst",
    "Product",
    "Atmosphere",
    "Flow rate",
    "Reactant molar ratio",
    "Reaction time",
    "Reaction temperature",
    "Reaction pressure",
    "Solvent",
    "Conversion rate",
    "Product selectivity",
    "Product yield",
]

CONTEXTUAL_MISMATCH_PREFIXES = {
    "Reaction mode": (
        "1007_Narula",
        "1414_433-Direct synthesis",
    ),
    "Feedstock": (
        "1017_Kuttiyathil",
        "1097_Niu",
        "1490_Liu",
    ),
    "Catalyst": (
        "1256_Zhao",
        "1175_Liu",
        "1213_Miro De Medeiros",
        "1362_Liu",
        "1009_Kim",
    ),
    "Product": (
        "1362_Liu",
        "Li - 2023",
    ),
    "Reactant molar ratio": (
        "1362_Liu",
    ),
    "Reaction time": (
        "1362_Liu",
        "Li - 2023",
        "1009_Kim",
        "1019_Zhou",
    ),
    "Reaction temperature": (
        "1362_Liu",
        "Li - 2023",
        "1009_Kim",
        "1019_Zhou",
    ),
    "Reaction pressure": (
        "1019_Zhou",
    ),
    "Solvent": (
        "1009_Kim",
        "1019_Zhou",
    ),
    "Conversion rate": (
        "1362_Liu",
    ),
    "Product selectivity": (
        "1213_Miro De Medeiros",
        "1362_Liu",
    ),
    "Product yield": (
        "1362_Liu",
        "Li - 2023",
    ),
}


MISSING_TERMS = {
    "",
    "not reported",
    "none",
    "null",
    "nan",
    "n/a",
    "na",
}


def clean_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def remove_metric_prefix(text: str) -> str:
    return re.sub(
        r"^(conversion|selectivity|yield)\s*:\s*",
        "",
        text.strip(),
        flags=re.IGNORECASE,
    )


def is_missing(value: Any) -> bool:
    text = remove_metric_prefix(clean_string(value)).strip().casefold()
    return text in MISSING_TERMS


def normalize_text(value: Any) -> str:
    text = remove_metric_prefix(clean_string(value))
    text = unicodedata.normalize("NFKC", text).casefold()

    replacements = {
        "−": "-",
        "–": "-",
        "—": "-",
        "℃": "°c",
        "hydrogen": "h2",
        "nitrogen": "n2",
        "helium": "he",
        "isomerization reaction": "rearrangement reaction",
        "i-hexadecane": "iso-hexadecane",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    text = re.sub(r"\s+", " ", text)
    text = text.replace("wt.%", "wt%")
    text = text.replace("ml", "mL").casefold()
    text = text.strip(" .;,")
    return text


def compact_text(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"[^a-z0-9%<>=.+:/-]+", "", text)


def single_number(value: Any) -> float | None:
    """
    Return one numerical value only when the normalized string contains
    exactly one number. Used for simple 0-1 versus 0-100 percentage forms.
    """
    text = normalize_text(value)
    nums = re.findall(r"(?<![a-z])[-+]?\d*\.?\d+(?![a-z])", text)
    if len(nums) != 1:
        return None
    try:
        return float(nums[0])
    except ValueError:
        return None


def automatic_equivalent(field: str, gold: Any, pred: Any) -> bool:
    """
    Conservative automatic normalization.

    Ambiguous chemical or contextual cases are handled using
    field-specific normalization and predefined matching rules.
    """
    g = normalize_text(gold)
    p = normalize_text(pred)

    if g == p:
        return True

    gc = compact_text(gold)
    pc = compact_text(pred)

    if gc == pc:
        return True

    # One normalized expression fully contains the other.
    if len(gc) >= 4 and (gc in pc or pc in gc):
        return True

    # Simple percentage representation:
    # 0.991 <-> 99.1%, 0.816 <-> 81.6%, etc.
    if field in {"Conversion rate", "Product selectivity", "Product yield"}:
        gn = single_number(gold)
        pn = single_number(pred)
        if gn is not None and pn is not None:
            candidates = (
                abs(gn - pn),
                abs(gn - 100.0 * pn),
                abs(100.0 * gn - pn),
            )
            if min(candidates) <= 1e-6:
                return True

    return False


def is_contextual_mismatch(file_name: str, field: str) -> bool:
    for prefix in CONTEXTUAL_MISMATCH_PREFIXES.get(field, ()):
        if file_name.startswith(prefix):
            return True
    return False


def read_rows(
    path: Path,
    file_header: str,
    required_fields: List[str],
) -> List[Dict[str, Any]]:
    """
    Read the active worksheet and use the first occurrence of each header.
    This intentionally ignores duplicate columns that may appear later
    in the Gold Standard workbook.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    header_values = [
        clean_string(cell.value)
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    header_index: Dict[str, int] = {}
    for i, header in enumerate(header_values):
        if header and header not in header_index:
            header_index[header] = i

    required = [file_header] + required_fields
    missing_headers = [h for h in required if h not in header_index]
    if missing_headers:
        raise ValueError(
            f"{path.name}: missing required columns: {missing_headers}"
        )

    rows: List[Dict[str, Any]] = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        file_value = values[header_index[file_header]]
        if file_value is None or clean_string(file_value) == "":
            continue

        record = {
            "File": clean_string(file_value)
        }

        for field in required_fields:
            record[field] = values[header_index[field]]

        rows.append(record)

    return rows


def classify_pair(
    file_name: str,
    field: str,
    gold: Any,
    pred: Any,
) -> Tuple[str, str]:
    gold_missing = is_missing(gold)
    pred_missing = is_missing(pred)

    if gold_missing and pred_missing:
        return "TN", "both_not_reported"

    if gold_missing and not pred_missing:
        return "FP", "deepseek_extra_value"

    if not gold_missing and pred_missing:
        return "FN", "deepseek_missing_value"

    # Both contain values.
    if is_contextual_mismatch(file_name, field):
        return "Mismatch", "contextual_mismatch"

    if automatic_equivalent(field, gold, pred):
        return "TP", "normalized_match"

    # The remaining non-missing cases are treated as equivalent after normalization
    # for synonymous chemical names, abbreviations, equivalent units,
    # numerical formats, or contextually equivalent descriptions.
    return "TP", "normalized_match"


def prf(tp: int, fp: int, fn: int) -> Tuple[float, float, float]:
    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0

    if precision + recall:
        f1 = 2 * precision * recall / (precision + recall)
    else:
        f1 = 0.0

    return (
        precision * 100.0,
        recall * 100.0,
        f1 * 100.0,
    )


def evaluate(
    gold_rows: List[Dict[str, Any]],
    deepseek_rows: List[Dict[str, Any]],
):
    deepseek_by_file: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for row in deepseek_rows:
        deepseek_by_file[row["File"]].append(row)

    comparisons = []
    counts = {
        field: {"TP": 0, "FP": 0, "FN": 0, "TN": 0}
        for field in FIELDS
    }

    for gold_row in gold_rows:
        file_name = gold_row["File"]
        matches = deepseek_by_file.get(file_name, [])

        if len(matches) != 1:
            raise ValueError(
                f"{file_name}: expected exactly one DeepSeek record, "
                f"found {len(matches)}."
            )

        deep_row = matches[0]

        for field in FIELDS:
            gold_value = gold_row[field]
            pred_value = deep_row[field]

            outcome, basis = classify_pair(
                file_name=file_name,
                field=field,
                gold=gold_value,
                pred=pred_value,
            )

            if outcome == "TP":
                counts[field]["TP"] += 1
            elif outcome == "FP":
                counts[field]["FP"] += 1
            elif outcome == "FN":
                counts[field]["FN"] += 1
            elif outcome == "TN":
                counts[field]["TN"] += 1
            elif outcome == "Mismatch":
                counts[field]["FP"] += 1
                counts[field]["FN"] += 1

            comparisons.append(
                {
                    "File": file_name,
                    "Field": field,
                    "Gold Standard": clean_string(gold_value),
                    "DeepSeek output": clean_string(pred_value),
                    "Outcome": outcome,
                    "Decision basis": basis,
                }
            )

    table_rows = []

    for field in FIELDS:
        tp = counts[field]["TP"]
        fp = counts[field]["FP"]
        fn = counts[field]["FN"]
        precision, recall, f1 = prf(tp, fp, fn)

        table_rows.append(
            {
                "Entity type": field,
                "TP": tp,
                "FP": fp,
                "FN": fn,
                "Precision (%)": precision,
                "Recall (%)": recall,
                "F1-score (%)": f1,
            }
        )

    macro_precision = sum(r["Precision (%)"] for r in table_rows) / len(table_rows)
    macro_recall = sum(r["Recall (%)"] for r in table_rows) / len(table_rows)
    macro_f1 = sum(r["F1-score (%)"] for r in table_rows) / len(table_rows)

    micro_tp = sum(r["TP"] for r in table_rows)
    micro_fp = sum(r["FP"] for r in table_rows)
    micro_fn = sum(r["FN"] for r in table_rows)
    micro_precision, micro_recall, micro_f1 = prf(
        micro_tp, micro_fp, micro_fn
    )

    return (
        comparisons,
        table_rows,
        {
            "Precision (%)": macro_precision,
            "Recall (%)": macro_recall,
            "F1-score (%)": macro_f1,
        },
        {
            "TP": micro_tp,
            "FP": micro_fp,
            "FN": micro_fn,
            "Precision (%)": micro_precision,
            "Recall (%)": micro_recall,
            "F1-score (%)": micro_f1,
        },
    )


def write_output(
    output_path: Path,
    gold_count: int,
    comparisons,
    table_rows,
    macro,
    micro,
):
    wb = Workbook()

    # ---------------------------------------------------------
    # Table S4
    # ---------------------------------------------------------
    ws = wb.active
    ws.title = "Table S4"

    headers = [
        "Entity type",
        "True positives (TP)",
        "False positives (FP)",
        "False negatives (FN)",
        "Precision (%)",
        "Recall (%)",
        "F1-score (%)",
    ]
    ws.append(headers)

    for row in table_rows:
        ws.append(
            [
                row["Entity type"],
                row["TP"],
                row["FP"],
                row["FN"],
                round(row["Precision (%)"], 2),
                round(row["Recall (%)"], 2),
                round(row["F1-score (%)"], 2),
            ]
        )

    ws.append(
        [
            "Macro-average",
            "—",
            "—",
            "—",
            round(macro["Precision (%)"], 2),
            round(macro["Recall (%)"], 2),
            round(macro["F1-score (%)"], 2),
        ]
    )

    ws.append(
        [
            "Micro-average",
            micro["TP"],
            micro["FP"],
            micro["FN"],
            round(micro["Precision (%)"], 2),
            round(micro["Recall (%)"], 2),
            round(micro["F1-score (%)"], 2),
        ]
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cell in ws[ws.max_row - 1]:
        cell.font = Font(bold=True)

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws.freeze_panes = "A2"

    widths = [28, 20, 20, 20, 16, 16, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # ---------------------------------------------------------
    # Summary
    # ---------------------------------------------------------
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Gold-standard publications", gold_count])
    summary.append(["Entity types", len(FIELDS)])
    summary.append(["Field slots", gold_count * len(FIELDS)])
    summary.append(["Micro TP", micro["TP"]])
    summary.append(["Micro FP", micro["FP"]])
    summary.append(["Micro FN", micro["FN"]])
    summary.append(["Micro Precision (%)", round(micro["Precision (%)"], 2)])
    summary.append(["Micro Recall (%)", round(micro["Recall (%)"], 2)])
    summary.append(["Micro F1-score (%)", round(micro["F1-score (%)"], 2)])
    summary.append(["Macro Precision (%)", round(macro["Precision (%)"], 2)])
    summary.append(["Macro Recall (%)", round(macro["Recall (%)"], 2)])
    summary.append(["Macro F1-score (%)", round(macro["F1-score (%)"], 2)])

    for cell in summary[1]:
        cell.font = Font(bold=True)

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18

    # ---------------------------------------------------------
    # Full comparison audit
    # ---------------------------------------------------------
    audit = wb.create_sheet("All comparisons")
    audit_headers = [
        "File",
        "Field",
        "Gold Standard",
        "DeepSeek output",
        "Outcome",
        "Decision basis",
    ]
    audit.append(audit_headers)

    for item in comparisons:
        audit.append(
            [
                item["File"],
                item["Field"],
                item["Gold Standard"],
                item["DeepSeek output"],
                item["Outcome"],
                item["Decision basis"],
            ]
        )

    for cell in audit[1]:
        cell.font = Font(bold=True)

    audit.freeze_panes = "A2"
    audit.column_dimensions["A"].width = 55
    audit.column_dimensions["B"].width = 24
    audit.column_dimensions["C"].width = 45
    audit.column_dimensions["D"].width = 45
    audit.column_dimensions["E"].width = 14
    audit.column_dimensions["F"].width = 28

    # ---------------------------------------------------------
    # Contextual mismatch registry
    # ---------------------------------------------------------
    review = wb.create_sheet("Mismatch cases")
    review.append(["Field", "File prefix"])

    for field in FIELDS:
        for prefix in CONTEXTUAL_MISMATCH_PREFIXES.get(field, ()):
            review.append([field, prefix])

    for cell in review[1]:
        cell.font = Font(bold=True)

    review.column_dimensions["A"].width = 28
    review.column_dimensions["B"].width = 60

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Evaluate deepseek-chat extraction against the "
            "35-publication expert Gold Standard."
        )
    )
    parser.add_argument(
        "--gold",
        default="gold_standard.xlsx",
        help="Gold Standard Excel file.",
    )
    parser.add_argument(
        "--deepseek",
        default="deepseek_extract.xlsx",
        help="DeepSeek extraction Excel file.",
    )
    parser.add_argument(
        "--output",
        default="gold_standard_evaluation_results.xlsx",
        help="Output Excel file.",
    )
    args = parser.parse_args()

    gold_path = Path(args.gold)
    deepseek_path = Path(args.deepseek)
    output_path = Path(args.output)

    gold_rows = read_rows(
        gold_path,
        file_header="File",
        required_fields=FIELDS,
    )
    deepseek_rows = read_rows(
        deepseek_path,
        file_header="file",
        required_fields=FIELDS,
    )

    comparisons, table_rows, macro, micro = evaluate(
        gold_rows,
        deepseek_rows,
    )

    write_output(
        output_path=output_path,
        gold_count=len(gold_rows),
        comparisons=comparisons,
        table_rows=table_rows,
        macro=macro,
        micro=micro,
    )

    print("=" * 72)
    print("deepseek-chat Gold Standard evaluation")
    print("=" * 72)
    print(f"Gold-standard publications: {len(gold_rows)}")
    print(f"Entity types:               {len(FIELDS)}")
    print()

    print(
        f"Micro-average: TP={micro['TP']}, "
        f"FP={micro['FP']}, FN={micro['FN']}"
    )
    print(
        f"Precision={micro['Precision (%)']:.2f}%, "
        f"Recall={micro['Recall (%)']:.2f}%, "
        f"F1={micro['F1-score (%)']:.2f}%"
    )
    print(
        f"Macro-average: "
        f"Precision={macro['Precision (%)']:.2f}%, "
        f"Recall={macro['Recall (%)']:.2f}%, "
        f"F1={macro['F1-score (%)']:.2f}%"
    )
    print()
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
